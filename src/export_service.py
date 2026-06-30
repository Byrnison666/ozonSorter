import os
from datetime import datetime
from typing import List
from sqlalchemy.orm import Session
from sqlalchemy import select, or_
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from .models import Shipment, AssignmentStatus, DeliveryPoint, ExportSession, ImportSession

class ExportService:
    def __init__(self, db_session: Session):
        self.session = db_session

    def generate_export(self, delivery_point: DeliveryPoint, output_path: str, import_session_id: int) -> ExportSession:
        import_session = self.session.get(ImportSession, import_session_id)
        if import_session is None:
            raise ValueError(f"Import session {import_session_id} not found")

        # Только посылки из текущего отчёта: импорт проставляет каждой встреченной
        # строке last_seen_import_session_id = id своей сессии. «Видели в этом отчёте»
        # = физически лежит на Казакова 68. Без этого условия TO_SHIP копится по всем
        # прошлым импортам и в файл попадают уже забранные с Донецка посылки.
        # Дедуп между отчётами: посылку, уже попавшую в выгрузку прошлой сессии,
        # повторно не показываем (она лежит на складе день за днём). Выгруженные
        # в ЭТОЙ же сессии пропускаем — чтобы повторная генерация того же отчёта
        # воспроизводила файл целиком, а не отдавала пустой.
        stmt = select(Shipment).where(
            Shipment.assigned_point == delivery_point,
            Shipment.assignment_status == AssignmentStatus.TO_SHIP,
            Shipment.last_seen_import_session_id == import_session_id,
            or_(
                Shipment.exported_import_session_id.is_(None),
                Shipment.exported_import_session_id == import_session_id,
            ),
        )
        shipments = self.session.execute(stmt).scalars().all()
        
        # Natural sort by cell
        shipments.sort(key=lambda s: self._natural_key(s.cell or ""))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime("%d,%m")
        
        # Styling
        damaged_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
        alignment = Alignment(wrap_text=True, vertical='top')
        
        for idx, shipment in enumerate(shipments, 1):
            # Column 1: описание товара — только название. Этикетка из отчёта Ozon
            # часто либо штрихкод (ii…), либо дубль номера отправления, поэтому в
            # описание не идёт. Фоллбэк — этикетка, затем сам номер (на пустое имя).
            description = (
                shipment.product_name
                or shipment.product_label
                or shipment.posting_number
            )
            ws.cell(row=idx, column=1, value=description).alignment = alignment

            # Column 2: номер отправления — первичный id посылки (в нём Ozon ID клиента)
            ws.cell(row=idx, column=2, value=shipment.posting_number).alignment = alignment
            
            # Column 3: cell
            ws.cell(row=idx, column=3, value=shipment.cell).alignment = alignment
            
            # Column 4: Damaged mark
            if shipment.is_damaged:
                ws.cell(row=idx, column=4, value="ПОВРЕЖДЕНО").alignment = alignment
                # Highlight row
                for col in range(1, 5):
                    ws.cell(row=idx, column=col).fill = damaged_fill

        # Column widths
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 14
        
        wb.save(output_path)
        
        # Отметить выгруженные — в отчёты следующих сессий они уже не попадут.
        for shipment in shipments:
            shipment.exported_import_session_id = import_session_id

        export_session = ExportSession(
            import_session_id=import_session_id,
            delivery_point=delivery_point,
            export_date=datetime.now(),
            file_path=output_path,
            shipments_count=len(shipments)
        )
        self.session.add(export_session)
        self.session.commit()
        return export_session

    def _natural_key(self, text: str) -> List:
        import re
        return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text)]
