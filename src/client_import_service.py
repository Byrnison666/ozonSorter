"""Массовый импорт клиентов из xlsx.

Формат: одна строка-заголовок (Ozon ID, ФИО, Телефон, Точка) и строки данных.
Невалидная строка не валит весь файл — собираем ошибки с номерами строк.
Существующий клиент (по Ozon ID) обновляется значениями из файла.
"""
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import Client, DeliveryPoint
from .parser import ExcelParser


# Логическое имя колонки -> допустимые заголовки (нормализованные: lower, без пробелов).
_HEADER_ALIASES = {
    "ozon_client_id": {"ozonid", "id", "озонid", "озонид"},
    "full_name": {"фио", "имя", "name"},
    "phone": {"телефон", "phone", "тел"},
    "point": {"точка", "точкаповыдаче", "точкапоумолчанию", "point"},
}

_POINT_ALIASES = {
    "комсомольская4": DeliveryPoint.KOMSOMOLSKAYA_4,
    "комсомольская": DeliveryPoint.KOMSOMOLSKAYA_4,
    "кольцевая16": DeliveryPoint.KOLTSEVAYA_16,
    "кольцевая": DeliveryPoint.KOLTSEVAYA_16,
}

_REQUIRED_COLUMNS = ("ozon_client_id", "point")

TEMPLATE_HEADERS = ["Ozon ID", "ФИО", "Телефон", "Точка"]


@dataclass
class ImportResult:
    added: int = 0
    updated: int = 0
    data_rows: int = 0
    errors: List[Tuple[int, str]] = field(default_factory=list)


def _norm(value) -> str:
    return str(value).strip().lower().replace(" ", "").replace("\n", "") if value is not None else ""


def _cell_to_str(value) -> str:
    """openpyxl может вернуть число для ID/телефона — приводим без .0."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


class ClientImportService:
    def __init__(self, db_session: Session):
        self.session = db_session

    @staticmethod
    def write_template(path: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Клиенты"
        ws.append(TEMPLATE_HEADERS)
        ws.append(["0224933356", "Иванов И.И.", "", "Комсомольская 4"])
        ws.append(["0301234567", "Петров П.П.", "", "Кольцевая 16"])
        widths = [16, 22, 16, 20]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        wb.save(path)

    def import_clients(self, file_path: str) -> ImportResult:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        header_row_idx, col_map = self._find_header(ws)
        if header_row_idx is None:
            raise ValueError(
                "Не найден заголовок. Обязательные колонки: «Ozon ID» и «Точка». "
                "Скачайте шаблон и заполните его."
            )

        # Существующие клиенты по нормализованному Ozon ID (без ведущих нулей).
        # Так «0224933356» из файла найдёт уже заведённого «224933356» — иначе
        # создавался бы второй клиент на одного человека, и матчинг посылок (он
        # тоже нормализует) молча отдавал бы посылку лишь одному из них.
        existing_by_norm: Dict[str, Client] = {}
        for c in self.session.execute(select(Client)).scalars().all():
            existing_by_norm.setdefault(ExcelParser.normalize_ozon_id(c.ozon_client_id), c)

        # Уже обработанные в этом файле id (нормализованные) → (строка, parsed).
        # Защита от дубля строки в одном файле: повтор с теми же данными молча
        # пропускаем, с другими (другая точка/ФИО) — сообщаем конфликт.
        seen_in_file: Dict[str, Tuple[int, Dict]] = {}

        result = ImportResult()
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            raw = {name: row[idx] for name, idx in col_map.items() if idx < len(row)}
            if not _cell_to_str(raw.get("ozon_client_id")):
                continue  # пустая строка — молча пропускаем

            result.data_rows += 1
            parsed, err = self._parse_row(raw)
            if err:
                result.errors.append((row_idx, err))
                continue

            norm = ExcelParser.normalize_ozon_id(parsed["ozon_client_id"])

            prev = seen_in_file.get(norm)
            if prev is not None:
                prev_idx, prev_parsed = prev
                if (prev_parsed["point"] != parsed["point"]
                        or prev_parsed["full_name"] != parsed["full_name"]
                        or prev_parsed["phone"] != parsed["phone"]):
                    result.errors.append((
                        row_idx,
                        f"Ozon ID {parsed['ozon_client_id']} уже в строке {prev_idx} "
                        f"с другими данными — конфликт, строка пропущена",
                    ))
                # одинаковый дубль строки — молча пропускаем
                continue
            seen_in_file[norm] = (row_idx, parsed)

            client = existing_by_norm.get(norm)
            if client is not None:
                # Канонизируем хранимый id (старые записи могли быть с нулём).
                client.ozon_client_id = norm
                client.full_name = parsed["full_name"]
                client.phone = parsed["phone"]
                client.fixed_delivery_point = parsed["point"]
                client.is_active = True
                result.updated += 1
            else:
                client = Client(
                    ozon_client_id=norm,
                    full_name=parsed["full_name"],
                    phone=parsed["phone"],
                    fixed_delivery_point=parsed["point"],
                )
                self.session.add(client)
                existing_by_norm[norm] = client
                result.added += 1

        self.session.commit()
        return result

    def _parse_row(self, raw: Dict) -> Tuple[Optional[Dict], Optional[str]]:
        ozon_id = _cell_to_str(raw.get("ozon_client_id"))
        if not ozon_id.isdigit():
            return None, f"Ozon ID «{ozon_id}» — только цифры, без дефисов и букв"

        point_raw = _norm(raw.get("point"))
        if not point_raw:
            return None, "Не указана точка («Комсомольская 4» или «Кольцевая 16»)"
        point = _POINT_ALIASES.get(point_raw)
        if point is None:
            return None, (
                f"Точка «{raw.get('point')}» — допустимо «Комсомольская 4» "
                f"или «Кольцевая 16»"
            )

        return {
            "ozon_client_id": ozon_id,
            "full_name": _cell_to_str(raw.get("full_name")) or None,
            "phone": _cell_to_str(raw.get("phone")) or None,
            "point": point,
        }, None

    def _find_header(self, ws) -> Tuple[Optional[int], Dict[str, int]]:
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            normalized = [_norm(c) for c in row]
            col_map: Dict[str, int] = {}
            for col_idx, cell in enumerate(normalized):
                if not cell:
                    continue
                for logical, aliases in _HEADER_ALIASES.items():
                    if cell in aliases and logical not in col_map:
                        col_map[logical] = col_idx
            if all(req in col_map for req in _REQUIRED_COLUMNS):
                return row_idx, col_map
        return None, {}
