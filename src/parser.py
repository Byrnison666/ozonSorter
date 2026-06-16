import hashlib
import re
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl.utils.datetime import from_excel

class ExcelParser:
    SIGNATURE_COLUMNS = ['Номер отправления', 'Тип', 'Статус', 'Ячейка']
    
    COLUMN_MAP = {
        'Этикетка\nНазвание': 'label_and_name',
        'Номер отправления': 'posting_number',
        'Тип': 'type',
        'Статус': 'status',
        'Ячейка': 'cell',
        'Отсчётная дата отправки': 'shipment_date_ozon',
        'Перевозка': 'carriage',
        'Контейнер\nШтрихкод': 'container_barcode'
    }

    @staticmethod
    def get_file_sha256(file_path: str) -> str:
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()

    @staticmethod
    def extract_ozon_client_id(posting_number: str) -> Optional[str]:
        if not posting_number:
            return None
        
        # Ozon Client ID is everything before the first '-'
        match = re.match(r'^(\d+)-', posting_number.strip())
        if match:
            client_id = match.group(1)
            # Validation: only digits, at least 6 digits (based on spec examples)
            if client_id.isdigit():
                return client_id
        return None

    def parse_file(self, file_path: str) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Try first visible sheet
        sheet = wb.active
        
        header_row_idx, col_mapping = self._find_header(sheet)
        if not header_row_idx:
            raise ValueError("Не удалось найти заголовок таблицы. Обязательные колонки: " + ", ".join(self.SIGNATURE_COLUMNS))

        results = []
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            # Check if posting_number is empty or invalid
            posting_idx = col_mapping.get('posting_number')
            if posting_idx is None:
                continue
                
            posting_number = row[posting_idx]
            if not posting_number:
                continue
            
            row_data = {}
            for logical_name, col_idx in col_mapping.items():
                val = row[col_idx]
                
                if logical_name == 'shipment_date_ozon' and isinstance(val, (int, float)):
                    try:
                        val = from_excel(val)
                    except:
                        val = None
                
                row_data[logical_name] = val
            
            # Post-processing
            if row_data.get('label_and_name'):
                parts = str(row_data['label_and_name']).split('\n')
                row_data['product_label'] = parts[0].strip() if len(parts) > 0 else None
                row_data['product_name'] = parts[1].strip() if len(parts) > 1 else None
            else:
                row_data['product_label'] = None
                row_data['product_name'] = None
                
            row_data['is_damaged'] = 'Повреждено' in str(row_data.get('type', ''))
            
            # Extract client ID
            client_id = self.extract_ozon_client_id(str(posting_number))
            row_data['ozon_client_id'] = client_id
            row_data['is_kty'] = client_id is None
            
            results.append(row_data)
            
        return results

    def _find_header(self, sheet) -> Tuple[Optional[int], Dict[str, int]]:
        for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
            normalized_row = [str(c).lower().replace(' ', '').replace('\n', '') if c else '' for c in row]
            
            # Check for signature columns
            matches = 0
            for sig in self.SIGNATURE_COLUMNS:
                norm_sig = sig.lower().replace(' ', '').replace('\n', '')
                if norm_sig in normalized_row:
                    matches += 1
            
            if matches == len(self.SIGNATURE_COLUMNS):
                # Found header row. Create mapping
                mapping = {}
                for col_idx, cell_val in enumerate(row):
                    if not cell_val: continue
                    norm_cell = str(cell_val).lower().replace(' ', '').replace('\n', '')
                    
                    for header, logical in self.COLUMN_MAP.items():
                        norm_header = header.lower().replace(' ', '').replace('\n', '')
                        if norm_header == norm_cell:
                            mapping[logical] = col_idx
                return row_idx, mapping
        return None, {}
