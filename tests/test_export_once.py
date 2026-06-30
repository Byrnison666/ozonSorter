"""Каждая посылка попадает в отчёт-отгрузку только один раз.

Отчёт Казакова — «остатки на складе»: посылка, пока её не забрали, висит в нём
день за днём. Без дедупликации она каждый день валилась бы в новую выгрузку.
Правило: выгрузили один раз → в следующих отчётах (новая сессия импорта) её уже
не показываем. Повторная генерация выгрузки того же импорта — воспроизводит файл
целиком (не «съедает» посылки).
"""
import os
import tempfile
import unittest

import openpyxl

from src.database import DatabaseManager
from src.models import Client, DeliveryPoint
from src.services import ImportService
from src.export_service import ExportService


def _report(postings):
    path = tempfile.mktemp(suffix=".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Этикетка\nНазвание", "Номер отправления", "Статус", "Ячейка"])
    for p in postings:
        ws.append([f"L\n{p}", p, "Готово к выдаче", "A-1"])
    wb.save(path)
    return path


class ExportOnceTest(unittest.TestCase):
    POINT = DeliveryPoint.KOMSOMOLSKAYA_4

    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.db = DatabaseManager(db_path=self.db_path)
        self.db.create_tables()
        self.session = self.db.get_session()
        self.session.add(Client(
            ozon_client_id="111", full_name="A",
            fixed_delivery_point=self.POINT,
        ))
        self.session.commit()
        self.imp = ImportService(self.session)
        self.exp = ExportService(self.session)
        self._files = []

    def tearDown(self):
        self.session.close()
        os.remove(self.db_path)
        for f in self._files:
            if os.path.exists(f):
                os.remove(f)

    def _import(self, postings):
        path = _report(postings)
        self._files.append(path)
        return self.imp.process_import(path)

    def _export(self, import_session_id):
        out = tempfile.mktemp(suffix=".xlsx")
        self._files.append(out)
        self.exp.generate_export(self.POINT, out, import_session_id)
        wb = openpyxl.load_workbook(out)
        return [r[1] for r in wb.active.iter_rows(values_only=True) if r[1]]

    def test_already_exported_parcel_skipped_next_day(self):
        d1 = self._import(["111-0001-1"])
        self.assertEqual(self._export(d1.id), ["111-0001-1"])

        # День 2: 0001 всё ещё на складе + новая 0002. В отчёт — только новая.
        d2 = self._import(["111-0001-1", "111-0002-1"])
        self.assertEqual(self._export(d2.id), ["111-0002-1"],
                         "уже выгруженная 0001 не дублируется")

    def test_regenerate_same_session_reproduces_file(self):
        d1 = self._import(["111-0001-1", "111-0002-1"])
        self.assertEqual(set(self._export(d1.id)), {"111-0001-1", "111-0002-1"})
        # Повторная генерация того же импорта — полный файл, не пустой.
        self.assertEqual(set(self._export(d1.id)), {"111-0001-1", "111-0002-1"},
                         "регенерация того же дня не съедает посылки")

    def test_returned_then_ready_again_re_exported(self):
        # Выгрузили → посылка ушла в возврат → снова «Готово к выдаче».
        # Мы её так и не забрали, значит обязана снова попасть в отгрузку.
        d1 = self._import(["111-0001-1"])
        self.assertEqual(self._export(d1.id), ["111-0001-1"])

        # День 2: возврат.
        path2 = tempfile.mktemp(suffix=".xlsx"); self._files.append(path2)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Этикетка\nНазвание", "Номер отправления", "Статус", "Ячейка"])
        ws.append(["L\n111-0001-1", "111-0001-1", "Отправить на склад", "A-1"])
        wb.save(path2)
        self.imp.process_import(path2)

        # День 3: снова готова к выдаче.
        d3 = self._import(["111-0001-1"])
        self.assertEqual(self._export(d3.id), ["111-0001-1"],
                         "вернувшаяся и снова готовая посылка должна выгрузиться")

    def test_export_count_reflects_dedup(self):
        d1 = self._import(["111-0001-1"])
        self._export(d1.id)
        d2 = self._import(["111-0001-1", "111-0002-1"])
        out = tempfile.mktemp(suffix=".xlsx")
        self._files.append(out)
        es = self.exp.generate_export(self.POINT, out, d2.id)
        self.assertEqual(es.shipments_count, 1, "в счётчик выгрузки только новая")


if __name__ == "__main__":
    unittest.main()
