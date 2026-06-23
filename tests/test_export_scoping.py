"""Регрессия: экспорт по точке берёт только текущий отчёт, а не накопленный TO_SHIP.

Баг: generate_export фильтровал по (assigned_point, TO_SHIP) без привязки к импорту,
поэтому в файл попадали посылки прошлых отчётов, которые уже забрали с Донецка
(статус TO_SHIP ничем не снимается). Фикс — отбор по last_seen_at >= started_at.
"""
import os
import tempfile
import unittest
from datetime import datetime

import openpyxl

from src.database import DatabaseManager
from src.models import (
    Client, Shipment, ImportSession,
    AssignmentStatus, DeliveryPoint,
)
from src.export_service import ExportService


def _shipment(session, posting, point, status, last_seen, import_session_id,
              last_seen_session_id=None, client_id=None):
    s = Shipment(
        posting_number=posting,
        client_id=client_id,
        ozon_client_id_raw=posting.split("-")[0],
        product_label="LBL",
        product_name="Товар",
        cell="A-01",
        assignment_status=status,
        assigned_point=point,
        import_session_id=import_session_id,
        last_seen_import_session_id=last_seen_session_id or import_session_id,
        first_seen_at=last_seen,
        last_seen_at=last_seen,
    )
    session.add(s)
    return s


class ExportScopingTest(unittest.TestCase):
    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.db = DatabaseManager(db_path=self.db_path)
        self.db.create_tables()
        self.session = self.db.get_session()
        self.export = ExportService(self.session)

        # День 1: отчёт в 10:00, посылка A (FIXED → TO_SHIP, Комсомольская).
        self.day1 = ImportSession(
            source_file_name="day1.xlsx", source_file_sha256="h1",
            started_at=datetime(2026, 6, 1, 10, 0, 0),
        )
        self.session.add(self.day1)
        self.session.flush()
        _shipment(
            self.session, "111-A-1", DeliveryPoint.KOMSOMOLSKAYA_4,
            AssignmentStatus.TO_SHIP, datetime(2026, 6, 1, 10, 0, 5), self.day1.id,
        )

        # День 2: отчёт в 10:00 следующего дня. A уже забрали — её НЕТ в отчёте,
        # last_seen_at остался вчерашним. Появилась B (TO_SHIP, Комсомольская).
        self.day2 = ImportSession(
            source_file_name="day2.xlsx", source_file_sha256="h2",
            started_at=datetime(2026, 6, 2, 10, 0, 0),
        )
        self.session.add(self.day2)
        self.session.flush()
        _shipment(
            self.session, "222-B-1", DeliveryPoint.KOMSOMOLSKAYA_4,
            AssignmentStatus.TO_SHIP, datetime(2026, 6, 2, 10, 0, 5), self.day2.id,
        )
        # D пришла в день 1 и всё ещё лежит на Донецке в день 2 (есть в обоих
        # отчётах). import_session_id остался день 1, но last_seen — день 2.
        _shipment(
            self.session, "444-D-1", DeliveryPoint.KOMSOMOLSKAYA_4,
            AssignmentStatus.TO_SHIP, datetime(2026, 6, 2, 10, 0, 6), self.day1.id,
            last_seen_session_id=self.day2.id,
        )
        self.session.commit()

    def tearDown(self):
        self.session.close()
        os.remove(self.db_path)

    def _export_postings(self, point, import_session_id):
        out = os.path.join(tempfile.gettempdir(), f"exp_{import_session_id}_{point.value}.xlsx")
        es = self.export.generate_export(point, out, import_session_id)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        postings = [row[1] for row in ws.iter_rows(values_only=True) if row[1]]
        os.remove(out)
        return es.shipments_count, postings

    def test_export_includes_only_current_report(self):
        # День 2 — новая B и всё ещё лежащая D. Забранную вчера A не показывает.
        count, postings = self._export_postings(DeliveryPoint.KOMSOMOLSKAYA_4, self.day2.id)
        self.assertEqual(count, 2)
        self.assertEqual(set(postings), {"222-B-1", "444-D-1"})
        self.assertNotIn("111-A-1", postings)

    def test_old_report_sees_only_what_it_last_saw(self):
        # День 1 — только A. D в день 1 ещё «не финальна» (last_seen = день 2),
        # поэтому в забор дня 1 не попадает.
        count, postings = self._export_postings(DeliveryPoint.KOMSOMOLSKAYA_4, self.day1.id)
        self.assertEqual(count, 1)
        self.assertEqual(postings, ["111-A-1"])

    def test_description_column_is_product_name(self):
        # Колонка A — описание = название товара (не этикетка/номер).
        out = os.path.join(tempfile.gettempdir(), "exp_desc.xlsx")
        self.export.generate_export(DeliveryPoint.KOMSOMOLSKAYA_4, out, self.day2.id)
        wb = openpyxl.load_workbook(out)
        descriptions = [row[0] for row in wb.active.iter_rows(values_only=True) if row[0]]
        os.remove(out)
        self.assertTrue(descriptions)
        self.assertTrue(all(d == "Товар" for d in descriptions),
                        f"описание должно быть названием товара, получено: {descriptions}")

    def test_unknown_import_session_raises(self):
        with self.assertRaises(ValueError):
            self.export.generate_export(DeliveryPoint.KOMSOMOLSKAYA_4, "x.xlsx", 99999)


if __name__ == "__main__":
    unittest.main()
