"""Фильтрация по колонке «Статус» отчёта склада.

Только «Готово к выдаче» можно забрать → идёт в отгрузку (TO_SHIP).
«Отправить на склад» / «Вернуть продавцу» — возврат: в отгрузку не идёт,
а уже стоявшая к отгрузке посылка снимается (RETURNED) и пропадает из экспорта.
Старый отчёт без колонки «Статус» — поведение прежнее (всё к отгрузке).
"""
import os
import tempfile
import unittest

import openpyxl
from sqlalchemy import select

from src.database import DatabaseManager
from src.models import Client, Shipment, DeliveryPoint, AssignmentStatus
from src.services import ImportService
from src.export_service import ExportService


def _make_report(rows, with_status=True):
    """rows: список (posting, status) или (posting,) для старого формата."""
    path = tempfile.mktemp(suffix=".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    if with_status:
        ws.append(["Этикетка\nНазвание", "Номер отправления", "Статус", "Ячейка"])
        for posting, status in rows:
            ws.append([f"LBL\n{posting}", posting, status, "A-01"])
    else:
        ws.append(["Этикетка\nНазвание", "Номер отправления", "Ячейка"])
        for (posting,) in rows:
            ws.append([f"LBL\n{posting}", posting, "A-01"])
    wb.save(path)
    return path


class StatusReadyTest(unittest.TestCase):
    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.db = DatabaseManager(db_path=self.db_path)
        self.db.create_tables()
        self.session = self.db.get_session()
        self.session.add(Client(
            ozon_client_id="111", full_name="Тест",
            fixed_delivery_point=DeliveryPoint.KOMSOMOLSKAYA_4,
        ))
        self.session.commit()
        self.importer = ImportService(self.session)
        self._files = []

    def tearDown(self):
        self.session.close()
        os.remove(self.db_path)
        for f in self._files:
            if os.path.exists(f):
                os.remove(f)

    def _import(self, rows, with_status=True):
        path = _make_report(rows, with_status=with_status)
        self._files.append(path)
        return self.importer.process_import(path)

    def _ship(self, posting):
        return self.session.execute(
            select(Shipment).where(Shipment.posting_number == posting)
        ).scalar_one()

    def _export_postings(self, import_session_id):
        out = tempfile.mktemp(suffix=".xlsx")
        self._files.append(out)
        ExportService(self.session).generate_export(
            DeliveryPoint.KOMSOMOLSKAYA_4, out, import_session_id
        )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        return {row[1] for row in ws.iter_rows(values_only=True) if row[1]}

    def test_ready_goes_to_ship(self):
        s = self._import([("111-0001-1", "Готово к выдаче")])
        self.assertEqual(self._ship("111-0001-1").assignment_status,
                         AssignmentStatus.TO_SHIP)
        self.assertEqual(s.new_to_ship_rows, 1)
        self.assertEqual(s.returned_rows, 0)
        self.assertIn("111-0001-1", self._export_postings(s.id))

    def test_return_status_not_shipped(self):
        for status in ("Отправить на склад", "Вернуть продавцу"):
            with self.subTest(status=status):
                posting = f"111-{status[:3]}-1"
                s = self._import([(posting, status)])
                self.assertEqual(self._ship(posting).assignment_status,
                                 AssignmentStatus.RETURNED)
                self.assertEqual(s.returned_rows, 1)
                self.assertNotIn(posting, self._export_postings(s.id))

    def test_ship_then_returned_is_pulled(self):
        posting = "111-0009-1"
        s1 = self._import([(posting, "Готово к выдаче")])
        self.assertEqual(self._ship(posting).assignment_status,
                         AssignmentStatus.TO_SHIP)
        # Следующий отчёт: та же посылка теперь возврат → снять с отгрузки.
        s2 = self._import([(posting, "Отправить на склад")])
        self.assertEqual(self._ship(posting).assignment_status,
                         AssignmentStatus.RETURNED)
        self.assertEqual(s2.returned_rows, 1)
        self.assertNotIn(posting, self._export_postings(s2.id))

    def test_returned_can_come_back(self):
        posting = "111-0010-1"
        self._import([(posting, "Вернуть продавцу")])
        self.assertEqual(self._ship(posting).assignment_status,
                         AssignmentStatus.RETURNED)
        s2 = self._import([(posting, "Готово к выдаче")])
        self.assertEqual(self._ship(posting).assignment_status,
                         AssignmentStatus.TO_SHIP)
        self.assertEqual(s2.new_to_ship_rows, 1)
        self.assertIn(posting, self._export_postings(s2.id))

    def test_old_format_without_status_ships(self):
        s = self._import([("111-0011-1",)], with_status=False)
        self.assertEqual(self._ship("111-0011-1").assignment_status,
                         AssignmentStatus.TO_SHIP)
        self.assertEqual(s.new_to_ship_rows, 1)


if __name__ == "__main__":
    unittest.main()
