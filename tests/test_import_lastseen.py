"""End-to-end: ImportService обновляет last_seen_import_session_id при повторной
встрече посылки. Так посылка, лежащая на Донецке несколько дней, остаётся в заборе
текущего отчёта, а не привязана навсегда к сессии первой встречи.
"""
import os
import tempfile
import unittest

import openpyxl
from sqlalchemy import select

from src.database import DatabaseManager
from src.models import Client, Shipment, DeliveryPoint, AssignmentStatus
from src.services import ImportService
from src.export_service import ExportService


def _make_report(postings):
    """Минимальный xlsx в формате отчёта Казакова 68."""
    path = tempfile.mktemp(suffix=".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Этикетка\nНазвание", "Номер отправления", "Тип", "Статус", "Ячейка"])
    for p in postings:
        ws.append([f"LBL\n{p}", p, "Обычный", "Готово к выдаче", "A-01"])
    wb.save(path)
    return path


class ImportLastSeenTest(unittest.TestCase):
    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.db = DatabaseManager(db_path=self.db_path)
        self.db.create_tables()
        self.session = self.db.get_session()
        # FIXED-клиент 111 → его посылки сразу TO_SHIP на Комсомольскую.
        self.session.add(Client(
            ozon_client_id="111", full_name="Тест",
            fixed_delivery_point=DeliveryPoint.KOMSOMOLSKAYA_4,
        ))
        self.session.commit()
        self.importer = ImportService(self.session)
        self.export = ExportService(self.session)
        self._files = []

    def tearDown(self):
        self.session.close()
        os.remove(self.db_path)
        for f in self._files:
            if os.path.exists(f):
                os.remove(f)

    def _import(self, postings):
        path = _make_report(postings)
        self._files.append(path)
        return self.importer.process_import(path)

    def test_reimport_moves_last_seen_to_current_session(self):
        day1 = self._import(["111-AAA-1"])
        ship = self.session.execute(
            select(Shipment).where(Shipment.posting_number == "111-AAA-1")
        ).scalar_one()
        self.assertEqual(ship.assignment_status, AssignmentStatus.TO_SHIP)
        self.assertEqual(ship.import_session_id, day1.id)
        self.assertEqual(ship.last_seen_import_session_id, day1.id)

        # Та же посылка в новом отчёте — всё ещё лежит на Донецке.
        day2 = self._import(["111-AAA-1"])
        self.session.refresh(ship)
        self.assertEqual(ship.import_session_id, day1.id, "первая встреча не меняется")
        self.assertEqual(ship.last_seen_import_session_id, day2.id, "last_seen едет вперёд")

        # Забор по дню 2 — видит посылку; по дню 1 — уже нет.
        out2 = tempfile.mktemp(suffix=".xlsx"); self._files.append(out2)
        es2 = self.export.generate_export(DeliveryPoint.KOMSOMOLSKAYA_4, out2, day2.id)
        self.assertEqual(es2.shipments_count, 1)

        out1 = tempfile.mktemp(suffix=".xlsx"); self._files.append(out1)
        es1 = self.export.generate_export(DeliveryPoint.KOMSOMOLSKAYA_4, out1, day1.id)
        self.assertEqual(es1.shipments_count, 0)

    def test_picked_up_shipment_drops_from_next_report(self):
        # День 1: посылки X и Y. День 2: X забрали — в отчёте только Y.
        day1 = self._import(["111-X-1", "111-Y-1"])
        day2 = self._import(["111-Y-1"])
        out = tempfile.mktemp(suffix=".xlsx"); self._files.append(out)
        es = self.export.generate_export(DeliveryPoint.KOMSOMOLSKAYA_4, out, day2.id)
        wb = openpyxl.load_workbook(out)
        postings = [r[1] for r in wb.active.iter_rows(values_only=True) if r[1]]
        self.assertEqual(postings, ["111-Y-1"])


if __name__ == "__main__":
    unittest.main()
